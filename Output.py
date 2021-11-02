# cython: language_level=3
# -*- coding: gbk -*-
import copy
import os
import secrets
import openpyxl
import Run
import Input

import datetime

def write_excel(name:str, datalist:list= [[]], sheet_num :int =0):
    '''
    :param name:The directory of the xlsx file
    '''
    workbook = openpyxl.Workbook()
    # # Create a new table
    if sheet_num<=1:
        sheet = workbook.create_sheet('demo',index=0)
        for row in range(len(datalist)):
            for column in range(len(datalist[row])):
                sheet.cell(row = row+1, column = column+1).value = str(datalist[row][column])
    if not os.path.exists(os.path.dirname(name)):# Create a directory if it does not exist
        os.makedirs(os.path.dirname(name))
    workbook.save(name)
    print('保存成功: ',name)

def kp_output(yuan,filename,data,result,correlation_x_show=[]):
    '''
    :param filename: File path
    :param data: Raw data
    :param yuan.Standard:Model selection criteria
    :param yuan.Standards:
    '''
    score_test=[]
    time_now=datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S.%f')##There is really the same time
    time_now=time_now[:-4]+('%04d'%secrets.randbelow(10000))
    score_all=[]
    filename_wen=filename[:-5]
    Standard_tong=[]

    ####Testing accuracy
    data1=copy.deepcopy(data)
    result1= Run.run_one_n(Type=yuan.Type, par=result['par'], data=data1, pretreat=yuan.pretreat, test_size=abs(yuan.test_size),
                           Random_State='no', Random_Fen=yuan.Random_Fen, use_y_clf='on', use_y_test='on', save_clf='on')
    for Standard1 in yuan.Standards:
        if Standard1==yuan.Standard:
            y_test=copy.deepcopy(result1['y_test'])
            y_clf=copy.deepcopy(result1['y_clf'])

    for Standard1 in yuan.Standards:
        Standard_tong.append(Standard1)
        if score_test !='off':
            score_test.append(Run.Calculate_Standard(Standard1,y_test,y_clf,negative='off'))


    filename1=os.path.basename(filename)
    tem= Input.read_excel(filename)[0][1:]
    comb_x_name=''
    for i in range(len(result['comb_x'])):
        comb_x_name=comb_x_name+str(tem[result['comb_x'][i]])+','
    comb_x_name=comb_x_name[:-1]

    show=[['Name','Type']]
    show_first=[[filename1,yuan.Type]]
    show[0]=show[0]+['Ranking']
    show_first[0]=show_first[0]+[yuan.Type_check]
    show[0]=show[0]+['Best_par']
    show_first[0]=show_first[0]+[result['par']]
    show[0]=show[0]+Standard_tong
    show_first[0]=show_first[0]+score_test
    show[0]=show[0]+['selected_var','var_sort','var_num','var_name']
    show_first[0]=show_first[0]+[result['comb_x'],str(correlation_x_show),str(len(result['comb_x'])),comb_x_name]

    show[0]=show[0]+['Mark']
    show_first[0]=show_first[0]+[time_now]

    show=show+show_first
    for i in range(len(score_all)):
        show.append(score_all[i]+[result['y_score'][i]['Random_state']])
    if len(y_test)<65500:
        show.append(['Measured','Predicted'])
        for j in range(len(y_test)):
            show.append([y_test[j][0],y_clf[j]])
    else:
        print('len(y_test):',len(y_test))

    filename = filename_wen+'/'+yuan.Type+'_'+time_now+'/'+yuan.Type+'_'+yuan.pretreat+'_'+yuan.Standard+'_'+time_now+'.xlsx'
    write_excel(filename,show)


if __name__ == '__main__':
    pass
